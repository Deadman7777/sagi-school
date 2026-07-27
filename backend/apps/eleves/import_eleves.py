"""Import d'élèves depuis un fichier Excel (.xlsx).

Le fichier suit le template généré par generer_template() : une ligne
d'en-têtes (l'ordre des colonnes est libre, seuls Nom complet et Section
sont obligatoires), puis une ligne par élève. analyser() ne touche pas à
la base : elle rend un rapport ligne par ligne que le front affiche avant
confirmation — la création effective se fait dans la vue, en transaction.

openpyxl est importé paresseusement : sur une installation Windows
existante où la lib manque, l'API répond un message clair au lieu de
faire tomber tout le module eleves.
"""
import datetime
import io
import re
import unicodedata


MAX_LIGNES = 2000

# clé interne -> en-tête affiché dans le template
COLONNES = {
    'nom_complet':      'Nom complet *',
    'genre':            'Genre (G/F)',
    'date_naissance':   'Date de naissance (JJ/MM/AAAA)',
    'lieu_naissance':   'Lieu de naissance',
    'section':          'Section *',
    'classe':           'Classe',
    'nom_pere':         'Nom du père',
    'telephone_pere':   'Téléphone père',
    'nom_mere':         'Nom de la mère',
    'telephone_mere':   'Téléphone mère',
    'nom_tuteur':       'Nom du tuteur',
    'telephone_tuteur': 'Téléphone tuteur',
    'lien_tuteur':      'Lien tuteur',
    'etat_sante':       'État de santé (Sain/Suivi/Chronique)',
    'observations_sante': 'Situation sanitaire',
    'date_inscription': "Date d'inscription (JJ/MM/AAAA)",
    'matricule':        'Matricule (vide = automatique)',
    # Situation réelle à la migration (calculée depuis les frais de la classe
    # + le prorata + le mois en cours) — prioritaire sur les colonnes détaillées
    'a_jour':           'À jour ce mois (O/N)',
    'dette_actuelle':   'Dette actuelle (montant)',
    # Ardoise des années d'AVANT (indépendante de l'année en cours, cumulable
    # avec n'importe laquelle des colonnes ci-dessus) — reliquat_migration
    'impaye_anterieur': 'Impayé antérieur (montant)',
    'origine_impaye':   'Origine impayé antérieur',
    # Reprise de soldes détaillée (migration) — voir apps.paiements.reprise
    'rep_inscription':  'Inscription déjà payée (O/N)',
    'rep_mensualites':  'Mensualités déjà payées (nombre)',
    'rep_uniforme':     'Uniforme déjà payé (O/N)',
    'rep_fournitures':  'Fournitures déjà payées (O/N)',
}

# en-têtes acceptés (normalisés) -> clé interne ; tolère les variantes
# courantes des fichiers que les écoles ont déjà
_SYNONYMES = {
    'nom complet':        'nom_complet',
    'nom et prenom':      'nom_complet',
    'prenom et nom':      'nom_complet',
    'nom':                'nom_complet',
    'genre':              'genre',
    'sexe':               'genre',
    'date de naissance':  'date_naissance',
    'date naissance':     'date_naissance',
    'ne le':              'date_naissance',
    'lieu de naissance':  'lieu_naissance',
    'lieu naissance':     'lieu_naissance',
    'section':            'section',
    'niveau':             'section',
    'classe':             'classe',
    'nom du pere':        'nom_pere',
    'pere':               'nom_pere',
    'telephone pere':     'telephone_pere',
    'tel pere':           'telephone_pere',
    'nom de la mere':     'nom_mere',
    'mere':               'nom_mere',
    'telephone mere':     'telephone_mere',
    'tel mere':           'telephone_mere',
    'nom du tuteur':      'nom_tuteur',
    'tuteur':             'nom_tuteur',
    'telephone tuteur':   'telephone_tuteur',
    'tel tuteur':         'telephone_tuteur',
    'lien tuteur':        'lien_tuteur',
    'lien de parente':    'lien_tuteur',
    'etat de sante':      'etat_sante',
    'sante':              'etat_sante',
    'situation sanitaire':'observations_sante',
    'observations sante': 'observations_sante',
    "date d'inscription": 'date_inscription',
    'date inscription':   'date_inscription',
    'matricule':          'matricule',
    'a jour':                    'a_jour',
    'a jour ce mois':            'a_jour',
    'ajour':                     'a_jour',
    'dette actuelle':            'dette_actuelle',
    'dette':                     'dette_actuelle',
    'dette en cours':            'dette_actuelle',
    'impaye anterieur':          'impaye_anterieur',
    'impayes anterieurs':        'impaye_anterieur',
    'dette anterieure':          'impaye_anterieur',
    'ancienne dette':            'impaye_anterieur',
    'reliquat':                  'impaye_anterieur',
    'reliquat anterieur':        'impaye_anterieur',
    'arriere':                   'impaye_anterieur',
    'arrieres':                  'impaye_anterieur',
    'ardoise':                   'impaye_anterieur',
    'origine impaye anterieur':  'origine_impaye',
    'origine impaye':            'origine_impaye',
    'annee impaye':              'origine_impaye',
    'inscription deja payee':    'rep_inscription',
    'inscription payee':         'rep_inscription',
    'mensualites deja payees':   'rep_mensualites',
    'mensualites payees':        'rep_mensualites',
    'mois payes':                'rep_mensualites',
    'uniforme deja paye':        'rep_uniforme',
    'uniforme paye':             'rep_uniforme',
    'fournitures deja payees':   'rep_fournitures',
    'fournitures payees':        'rep_fournitures',
}

_FORMATS_DATE = ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y')

# Noms de mois FR (normalisés, sans accent) pour les dates « mois-année »
# saisies quand le jour exact n'est pas connu (ex. « Juillet 2025 »).
_MOIS_NOMS = {
    'janvier': 1, 'jan': 1, 'fevrier': 2, 'fev': 2, 'mars': 3,
    'avril': 4, 'avr': 4, 'mai': 5, 'juin': 6, 'juillet': 7, 'juil': 7,
    'aout': 8, 'septembre': 9, 'sept': 9, 'sep': 9, 'octobre': 10, 'oct': 10,
    'novembre': 11, 'nov': 11, 'decembre': 12, 'dec': 12,
}


def _mois_annee(val):
    """« Juillet 2025 », « 07/2025 », « 2025-07 » → date au 1er du mois, ou None.
    Sert quand le jour d'inscription n'est pas connu (le prorata n'utilise que
    le mois et l'année)."""
    txt = _norm(val)
    if not txt:
        return None
    m = re.match(r'^([a-z]+)\.?\s+(\d{4})$', txt)          # « juillet 2025 »
    if m and m.group(1) in _MOIS_NOMS:
        return datetime.date(int(m.group(2)), _MOIS_NOMS[m.group(1)], 1)
    m = re.match(r'^(\d{1,2})[/\-.](\d{4})$', txt)         # « 07/2025 »
    if m and 1 <= int(m.group(1)) <= 12:
        return datetime.date(int(m.group(2)), int(m.group(1)), 1)
    m = re.match(r'^(\d{4})[/\-.](\d{1,2})$', txt)         # « 2025-07 »
    if m and 1 <= int(m.group(2)) <= 12:
        return datetime.date(int(m.group(1)), int(m.group(2)), 1)
    return None


def _openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError(
            "La bibliothèque openpyxl n'est pas installée sur ce poste. "
            "Exécutez « python -m pip install openpyxl » puis relancez l'application."
        )


def _norm(s):
    """minuscules, sans accents, espaces réduits — pour comparer noms et en-têtes."""
    if s is None:
        return ''
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.lower().split())


def _norm_entete(s):
    """Normalise un en-tête : on coupe la partie entre parenthèses et les astérisques."""
    s = str(s or '').split('(')[0].replace('*', '')
    return _norm(s)


def _texte(val):
    """Cellule -> str propre. Les téléphones/matricules saisis en numérique
    dans Excel arrivent en float (771234567.0) : on retombe sur l'entier."""
    if val is None:
        return ''
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _date(val):
    """Cellule -> (date | None, erreur | None). Accepte les vraies dates Excel
    et les chaînes aux formats usuels."""
    if val is None or val == '':
        return None, None
    if isinstance(val, datetime.datetime):
        return val.date(), None
    if isinstance(val, datetime.date):
        return val, None
    txt = str(val).strip()
    for fmt in _FORMATS_DATE:
        try:
            return datetime.datetime.strptime(txt, fmt).date(), None
        except ValueError:
            continue
    return None, f"date illisible « {txt} » (attendu JJ/MM/AAAA)"


def _oui_non(val):
    """Cellule O/N -> (bool, avertissement | None). Vide = Non."""
    n = _norm(val)
    if not n or n in ('n', 'non', 'no', '0', 'faux', 'false'):
        return False, None
    if n in ('o', 'oui', 'y', 'yes', 'x', '1', 'vrai', 'true'):
        return True, None
    return False, f"valeur « {val} » non reconnue (attendu O ou N) — considérée comme Non"


def _entier(val):
    """Cellule -> (int, erreur | None). Vide = 0."""
    if val is None or val == '':
        return 0, None
    try:
        n = int(float(val))
        if n < 0:
            return 0, f"nombre négatif « {val} » — ramené à 0"
        return n, None
    except (TypeError, ValueError):
        return 0, f"nombre illisible « {val} » — considéré comme 0"


def _montant(val):
    """Cellule -> (float >= 0, avertissement | None). Vide = 0.0.
    Tolère « 30 000 », « 30.000 » (séparateur de milliers) ; FCFA sans décimale."""
    if val is None or val == '':
        return 0.0, None
    if isinstance(val, (int, float)):
        return (float(val), None) if val >= 0 else (0.0, f"montant négatif « {val} » — ramené à 0")
    txt = (str(val).strip()
           .replace(' ', '').replace(' ', '').replace('\xa0', '')
           .replace('.', '').replace(',', '.'))
    try:
        n = float(txt)
    except ValueError:
        return 0.0, f"montant illisible « {val} » — ignoré"
    return (n, None) if n >= 0 else (0.0, f"montant négatif « {val} » — ramené à 0")


def _tel(val):
    """Nettoie un téléphone : si plusieurs numéros (séparés par / ou ;),
    garde le premier ; tronque à 20 caractères. Rend (str, avertissement|None)."""
    t = _texte(val)
    if not t:
        return '', None
    warn = None
    for sep in ('/', ';'):
        if sep in t:
            t = t.split(sep)[0].strip()
            warn = 'plusieurs numéros — seul le premier est conservé'
            break
    if len(t) > 20:
        t = t[:20]
        warn = 'numéro trop long — tronqué'
    return t, warn


def _mois_echus_import(exercice, date_insc, today):
    """Mensualités échues à ce jour (mois courant inclus), plafonnées au nombre
    dû au prorata de l'entrée. Réplique Eleve.mois_echus pour l'import (les
    élèves importés sont en régime EXERCICE)."""
    debut = exercice.date_debut
    insc = date_insc or debut
    mois_avant = max(0, (insc.year - debut.year) * 12 + (insc.month - debut.month)) if insc > debut else 0
    nb_dues = max(0, exercice.nb_mensualites - mois_avant)
    elapsed_incl = (today.year - debut.year) * 12 + (today.month - debut.month) + 1
    return max(0, min(elapsed_incl - mois_avant, nb_dues))


def _genre(val):
    """-> ('G'|'F'|'', avertissement | None)"""
    n = _norm(val)
    if not n:
        return '', None
    if n in ('g', 'm', 'h', 'garcon', 'masculin', 'homme'):
        return 'G', None
    if n in ('f', 'fille', 'feminin', 'femme'):
        return 'F', None
    return '', f"genre « {val} » non reconnu (attendu G ou F) — laissé vide"


def _etat_sante(val):
    """-> ('SAIN'|'SUIVI'|'CHRONIQUE', avertissement | None). Vide = Sain."""
    n = _norm(val)
    if not n or n in ('sain', 'saine', 'ok', 'bon', 'bonne sante', 'rien', 'ras'):
        return 'SAIN', None
    if n in ('suivi', 'sous suivi', 'suivi medical', 'sous suivi medical', 'a suivre'):
        return 'SUIVI', None
    if n in ('chronique', 'maladie chronique', 'malade', 'maladie'):
        return 'CHRONIQUE', None
    return 'SAIN', f"état de santé « {val} » non reconnu (Sain/Suivi/Chronique) — considéré Sain"


def generer_template(tenant):
    """Construit le classeur template et le rend en bytes (BytesIO)."""
    openpyxl = _openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from .models import Section

    wb = openpyxl.Workbook()

    # ── Feuille de saisie ────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Élèves'
    entetes = list(COLONNES.values())
    ws.append(entetes)
    fill = PatternFill('solid', fgColor='00B894')
    for col, _ in enumerate(entetes, start=1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = fill
        c.alignment = Alignment(vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = max(len(entetes[col - 1]) + 2, 16)
    ws.freeze_panes = 'A2'

    # ── Feuille consignes ────────────────────────────────────────────────
    sections = list(Section.objects.filter(tenant=tenant).values_list('nom', flat=True))
    aide = wb.create_sheet('Consignes')
    aide.column_dimensions['A'].width = 46
    aide.column_dimensions['B'].width = 64
    lignes = [
        ('Consignes de remplissage', ''),
        ('', ''),
        ('Nom complet *', 'Obligatoire. Ex. : Moussa Ndiaye'),
        ('Genre (G/F)', 'G ou F (Garçon / Fille). Optionnel.'),
        ('Date de naissance', 'JJ/MM/AAAA ou format date Excel. Optionnel.'),
        ('Section *', 'Obligatoire. Doit exister dans SAGI SCHOOL (liste ci-dessous).'),
        ('Classe', "Optionnel. Nom exact de la classe (ex. CI A) si l'école en utilise."),
        ('Téléphones', 'Chiffres uniquement, ex. 771234567. Optionnel.'),
        ('Tuteur', "Optionnel. À renseigner si le tuteur diffère des parents (nom, téléphone, lien)."),
        ('État de santé', 'Optionnel. Sain, Suivi (sous suivi médical) ou Chronique. Vide = Sain.'),
        ('Situation sanitaire', 'Optionnel. Allergies, maladies, traitements en cours.'),
        ("Date d'inscription", "Optionnel. Vide = début de l'année scolaire (aucun prorata)."),
        ('Matricule', 'Laissez VIDE pour une génération automatique (recommandé).'),
        ('', ''),
        ('Situation à la migration (le plus simple)', 'Deux colonnes suffisent pour refléter '
         "l'état réel de l'élève, calculé depuis les frais de sa classe et le mois en cours :"),
        ('À jour ce mois (O/N)', "O = l'élève a tout réglé jusqu'au mois en cours (aucun arriéré)."),
        ('Dette actuelle (montant)', "Montant exactement dû à ce jour, ex. 30000. Le payé est "
         "reconstitué automatiquement. Prioritaire sur les colonnes détaillées ci-dessous."),
        ('', ''),
        ("Impayé des années d'AVANT (optionnel)", "Ce que la famille devait déjà en arrivant "
         "dans SAGI SCHOOL, toutes années confondues."),
        ('Impayé antérieur (montant)', "Montant global, ex. 45000. Aucun détail à fournir : "
         "un seul chiffre suffit. Il s'ajoute au dû de l'année en cours et se règle à part."),
        ('Origine impayé antérieur', "Optionnel. Texte libre : « 2024-2025 », « ardoise cahier », "
         "« ancien logiciel »… Sert seulement à vous souvenir d'où vient la dette."),
        ('', "Cette colonne est INDÉPENDANTE des précédentes : vous pouvez la remplir seule, "
         "ou en plus de « À jour » / « Dette actuelle ». Et vous pouvez la laisser vide "
         "aujourd'hui et la saisir plus tard, élève par élève, sans refaire l'import."),
        ('', ''),
        ('Reprise détaillée (optionnelle)', 'À la place, si vous préférez détailler ce qui a déjà'
         ' été encaissé AVANT SAGI SCHOOL. Ignorée si « À jour » / « Dette actuelle » est renseigné.'),
        ('Inscription / Uniforme / Fournitures déjà payés', 'O si déjà réglé, N ou vide sinon.'),
        ('Mensualités déjà payées', 'Nombre de mois déjà réglés depuis le début de l\'année'
         ' (ex. 3 = octobre, novembre, décembre payés). Vide = 0.'),
        ('', 'Les montants sont calculés à partir des frais de la section, enregistrés en'
         ' « Reprise » : le reste à payer et le suivi mensuel repartent d\'un état juste,'
         ' sans toucher à votre caisse.'),
        ('', ''),
        ('Sections existantes dans votre école :', ', '.join(sections) or
         "(aucune — créez d'abord vos sections dans SAGI SCHOOL)"),
        ('', ''),
        ('Limite', f'{MAX_LIGNES} élèves par fichier maximum.'),
    ]
    for a, b in lignes:
        aide.append([a, b])
    aide['A1'].font = Font(bold=True, size=13)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def analyser(fichier, tenant, exercice):
    """Lit le .xlsx et rend le rapport {'resume': {...}, 'lignes': [...]}.

    Chaque ligne : {ligne, nom_complet, section, statut OK|DOUBLON|ERREUR,
    erreurs[], avertissements[], data{...}} — data contient les valeurs
    nettoyées prêtes pour Eleve.objects.create().
    Lève ValueError (fichier illisible / en-têtes absents / trop de lignes).
    """
    openpyxl = _openpyxl()
    from .models import Eleve, Section
    from apps.academique.models import Classe

    try:
        wb = openpyxl.load_workbook(fichier, read_only=True, data_only=True)
    except Exception:
        raise ValueError("Fichier illisible — envoyez un .xlsx (Excel 2007+), pas un .xls ni un CSV.")

    ws = wb['Élèves'] if 'Élèves' in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)

    # 1re ligne non vide = en-têtes
    entetes = None
    for row in rows:
        if any(v not in (None, '') for v in row):
            entetes = row
            break
    if entetes is None:
        raise ValueError('Fichier vide.')

    colmap = {}   # index -> clé interne
    for i, h in enumerate(entetes):
        cle = _SYNONYMES.get(_norm_entete(h))
        if cle and cle not in colmap.values():
            colmap[i] = cle
    if 'nom_complet' not in colmap.values() or 'section' not in colmap.values():
        raise ValueError(
            'En-têtes non reconnus — il faut au minimum les colonnes '
            '« Nom complet » et « Section ». Téléchargez le template fourni.'
        )

    sections = {_norm(s.nom): s for s in Section.objects.filter(tenant=tenant)}
    classes  = {_norm(c.nom): c for c in Classe.objects.filter(tenant=tenant)}
    deja_la  = set(
        (_norm(n), d) for n, d in
        Eleve.objects.filter(tenant=tenant, exercice=exercice)
                     .values_list('nom_complet', 'date_naissance')
    )
    matricules_pris = set(
        m for m in Eleve.objects.filter(tenant=tenant).values_list('matricule', flat=True) if m
    )

    lignes, vus_fichier = [], set()
    today = datetime.date.today()   # référence du mois courant pour « à jour »/dette
    # Tailles max des champs texte → garde-fou anti « value too long » (500)
    _maxlen = {f.name: f.max_length for f in Eleve._meta.get_fields()
               if getattr(f, 'max_length', None)}
    no_ligne = 1  # la ligne d'en-têtes ; les données commencent après
    for row in rows:
        no_ligne += 1
        if not any(v not in (None, '') for v in row):
            continue
        if len(lignes) >= MAX_LIGNES:
            raise ValueError(f'Fichier trop volumineux : maximum {MAX_LIGNES} élèves par import.')

        brut = {cle: row[i] if i < len(row) else None for i, cle in colmap.items()}
        erreurs, avert = [], []

        nom = _texte(brut.get('nom_complet'))
        if not nom:
            erreurs.append('Nom complet manquant')

        genre, warn = _genre(brut.get('genre'))
        if warn:
            avert.append(warn)

        date_naiss, err = _date(brut.get('date_naissance'))
        if err:
            erreurs.append(f'Date de naissance : {err}')

        date_insc, err = _date(brut.get('date_inscription'))
        jour_estime = False
        if date_insc is None and brut.get('date_inscription') not in (None, ''):
            # jour non connu ? tolérer « Juillet 2025 » / « 07/2025 »
            ma = _mois_annee(brut.get('date_inscription'))
            if ma:
                date_insc, jour_estime, err = ma, True, None
                avert.append("Date d'inscription : jour non précisé — 1er du mois "
                             "retenu (affiché « mois année »)")
        if err:
            erreurs.append(f"Date d'inscription : {err}")
        if date_insc is None:
            # vide = présent depuis le début de l'année -> pas de prorata
            date_insc = exercice.date_debut

        nom_section = _texte(brut.get('section'))
        section = sections.get(_norm(nom_section))
        if not nom_section:
            erreurs.append('Section manquante')
        elif not section:
            erreurs.append(
                f'Section « {nom_section} » introuvable — créez-la dans '
                'Paramètres > Sections puis relancez l\'import'
            )

        nom_classe = _texte(brut.get('classe'))
        classe = classes.get(_norm(nom_classe)) if nom_classe else None
        if nom_classe and not classe:
            avert.append(f'Classe « {nom_classe} » inconnue — élève importé sans classe')

        etat_sante, warn_sante = _etat_sante(brut.get('etat_sante'))
        if warn_sante:
            avert.append(warn_sante)

        matricule = _texte(brut.get('matricule'))
        if matricule:
            if matricule in matricules_pris:
                erreurs.append(f'Matricule « {matricule} » déjà utilisé')
            else:
                matricules_pris.add(matricule)

        # ── Situation à la migration ─────────────────────────────────────
        # Deux façons, par ordre de priorité :
        #   1. « Dette actuelle » (montant) ou « À jour » (O) → reconstruit le
        #      payé à ce jour = dû échu − dette, depuis les frais de la classe,
        #      le prorata d'entrée et le mois en cours (dette 0 = à jour) ;
        #   2. colonnes détaillées « déjà payé » (inscription/mensualités/…).
        rep_inscription, w1 = _oui_non(brut.get('rep_inscription'))
        rep_uniforme,    w2 = _oui_non(brut.get('rep_uniforme'))
        rep_fournitures, w3 = _oui_non(brut.get('rep_fournitures'))
        rep_mensualites, w4 = _entier(brut.get('rep_mensualites'))
        for w, col in ((w1, 'Inscription déjà payée'), (w2, 'Uniforme déjà payé'),
                       (w3, 'Fournitures déjà payées'), (w4, 'Mensualités déjà payées')):
            if w:
                avert.append(f'{col} : {w}')
        if rep_mensualites > exercice.nb_mensualites:
            avert.append(f'Mensualités déjà payées : {rep_mensualites} > '
                         f'{exercice.nb_mensualites} mensualités de l\'exercice — plafonné')
            rep_mensualites = exercice.nb_mensualites

        a_jour, _wj = _oui_non(brut.get('a_jour'))
        dette_val, wd = _montant(brut.get('dette_actuelle'))
        dette_fournie = brut.get('dette_actuelle') not in (None, '')
        a_jour_fourni = brut.get('a_jour') not in (None, '')
        if wd:
            avert.append(f'Dette actuelle : {wd}')
        detaille_fourni = rep_inscription or rep_uniforme or rep_fournitures or rep_mensualites > 0

        montant_reprise = 0.0
        reprise_payload = {'inscription': rep_inscription, 'nb_mensualites': rep_mensualites,
                           'uniforme': rep_uniforme, 'fournitures': rep_fournitures}

        if section is not None and (dette_fournie or a_jour):
            if dette_fournie and a_jour_fourni:
                avert.append('« À jour » et « Dette actuelle » renseignés — '
                             '« Dette actuelle » prioritaire')
            if detaille_fourni:
                avert.append('Colonnes « déjà payé » ignorées au profit de '
                             '« À jour » / « Dette actuelle »')
            me = _mois_echus_import(exercice, date_insc, today)
            fi = float(section.frais_inscription); fm = float(section.frais_mensualite)
            fu = float(section.frais_uniforme);    ff = float(section.frais_fournitures)
            du = fi + fu + ff + fm * me
            dette = dette_val if dette_fournie else 0.0
            if dette > du:
                avert.append(f'Dette actuelle ({dette:.0f}) supérieure au dû à ce jour '
                             f'({du:.0f}) — plafonnée')
                dette = du
            paye = du - dette
            # Répartition : ponctuels réglés d'abord, la dette reste en arriérés
            # de mensualités (c'est ce que lisent les alertes).
            reste = paye
            m_i = min(reste, fi); reste -= m_i
            m_u = min(reste, fu); reste -= m_u
            m_f = min(reste, ff); reste -= m_f
            m_m = min(reste, fm * me)
            reprise_payload = {'montants': {
                'montant_inscription': round(m_i, 2), 'montant_mensualite': round(m_m, 2),
                'montant_uniforme': round(m_u, 2), 'montant_fournitures': round(m_f, 2),
            }}
            montant_reprise = round(paye, 2)
        elif section is not None:
            montant_reprise = float(
                (section.frais_inscription if rep_inscription else 0)
                + section.frais_mensualite * rep_mensualites
                + (section.frais_uniforme if rep_uniforme else 0)
                + (section.frais_fournitures if rep_fournitures else 0)
            )

        # ── Ardoise des années d'avant ───────────────────────────────────
        # Totalement indépendante de la situation de l'année en cours : une
        # école peut ne remplir QUE cette colonne, ou la cumuler avec « Dette
        # actuelle ». Le montant n'a pas à être justifié poste par poste —
        # c'est précisément ce qu'on n'obtient pas sur le terrain.
        # _montant ramène un négatif à 0 avec un avertissement : une faute de
        # frappe n'a pas à faire échouer la ligne entière (l'élève, lui, est bon).
        impaye_ant, wa = _montant(brut.get('impaye_anterieur'))
        if wa:
            avert.append(f'Impayé antérieur : {wa}')
        origine_impaye = _texte(brut.get('origine_impaye'))[:120]

        statut = 'ERREUR' if erreurs else 'OK'
        cle_doublon = (_norm(nom), date_naiss)
        if statut == 'OK':
            if cle_doublon in deja_la:
                statut = 'DOUBLON'
                avert.append('Déjà inscrit cette année (même nom et date de naissance) — ignoré')
            elif cle_doublon in vus_fichier:
                statut = 'DOUBLON'
                avert.append('Présent en double dans le fichier — ignoré')
            else:
                vus_fichier.add(cle_doublon)

        # Téléphones : plusieurs numéros « / » fréquents → garder le premier
        tel_pere,   wtp = _tel(brut.get('telephone_pere'))
        tel_mere,   wtm = _tel(brut.get('telephone_mere'))
        tel_tuteur, wtt = _tel(brut.get('telephone_tuteur'))
        for w, lbl in ((wtp, 'Téléphone père'), (wtm, 'Téléphone mère'),
                       (wtt, 'Téléphone tuteur')):
            if w:
                avert.append(f'{lbl} : {w}')

        data = {
            'nom_complet':      nom,
            'genre':            genre,
            'date_naissance':   date_naiss,   # DRF sérialise les dates en ISO dans le rapport
            'lieu_naissance':   _texte(brut.get('lieu_naissance')),
            'section_id':       section.id if section else None,
            'classe_id':        classe.id if classe else None,
            'nom_pere':         _texte(brut.get('nom_pere')),
            'telephone_pere':   tel_pere,
            'nom_mere':         _texte(brut.get('nom_mere')),
            'telephone_mere':   tel_mere,
            'nom_tuteur':       _texte(brut.get('nom_tuteur')),
            'telephone_tuteur': tel_tuteur,
            'lien_tuteur':      _texte(brut.get('lien_tuteur')),
            'etat_sante':       etat_sante,
            'observations_sante': _texte(brut.get('observations_sante')),
            'date_inscription': date_insc,
            'date_inscription_jour_estime': jour_estime,
            'matricule':        matricule or None,
        }
        # Garde-fou générique : tronquer toute valeur texte qui dépasse la
        # taille de son champ (un import ne doit jamais faire un 500).
        for champ, val in list(data.items()):
            ml = _maxlen.get(champ)
            if ml and isinstance(val, str) and len(val) > ml:
                data[champ] = val[:ml]
                avert.append(f'« {champ} » trop long ({len(val)}>{ml}) — tronqué')

        lignes.append({
            'ligne':          no_ligne,
            'nom_complet':    nom,
            'section':        section.nom if section else nom_section,
            'statut':         statut,
            'erreurs':        erreurs,
            'avertissements': avert,
            'montant_reprise': montant_reprise,
            'reprise': reprise_payload,
            'impaye_anterieur': round(impaye_ant, 2),
            'origine_impaye':   origine_impaye,
            'data': data,
        })

    ok = [l for l in lignes if l['statut'] == 'OK']
    resume = {
        'total':    len(lignes),
        'ok':       len(ok),
        'doublons': sum(1 for l in lignes if l['statut'] == 'DOUBLON'),
        'erreurs':  sum(1 for l in lignes if l['statut'] == 'ERREUR'),
        'reprises': sum(1 for l in ok if l['montant_reprise'] > 0),
        'montant_reprise': sum(l['montant_reprise'] for l in ok),
        'impayes_anterieurs':        sum(1 for l in ok if l['impaye_anterieur'] > 0),
        'montant_impaye_anterieur':  round(sum(l['impaye_anterieur'] for l in ok), 2),
    }
    return {'resume': resume, 'lignes': lignes}
