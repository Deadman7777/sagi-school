"""Reconstruit le « déjà payé » / reste dû réel par élève (Complexe Shoumoul, 2026).

Modèle validé avec le directeur :
  - À jour (col « À jour ce mois » = O) → reste = 5 mois (août→déc) ×
    (mensualité + services de l'élève) + renouvellement dû.
  - Non à jour (N) → reste = sa « Dette actuelle » (elle vaut le reste complet).
  - Renouvellement 55 000/an (remplace l'inscription) ; versé selon la liste.
  - 3 nouveaux paient l'inscription 185 000 (au lieu du renouvellement) :
    Goundo Momi Keita (soldé), Aïssatou Keïta (soldé), Mamy Daya CISSOKHO
    (reste 85 000).

Le reste voulu est atteint en réglant le « déjà payé » de la reprise à
`total_attendu − reste`. 706 (13 410 500) inchangé : les reprises reconstruites
sont neutralisées en 890. Le reste dû par élève devient juste (objectif
prioritaire) ; le déjà payé ne colle pas au cash (journal de caisse incomplet).

  python manage.py recaler_reste_du_shoumoul --fichier "C:\\...\\import_eleves_sagi_NV.xlsx"
      [--tenant_id UUID] [--exercice 2026] [--appliquer]

Sans --appliquer : rapport par élève + alertes. Idempotent (remet à plat les
reprises existantes avant de reconstruire).
"""
import unicodedata
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Sum

from apps.tenants.models import Tenant
from apps.paiements.models import Exercice, Paiement
from apps.comptabilite.models import JournalEntry
from apps.paiements.reprise import creer_paiement_reprise

RENOUVELLEMENT = Decimal('55000')
MOIS_PAYES = 7            # jan→juil
MOIS_RESTANTS = 5         # août→déc


def _norm(s):
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode()
    return ' '.join(s.lower().split())


# Renouvellement versé (noms corrigés d'après la base).
RENOUV_VERSE = {_norm(n): Decimal(str(v)) for n, v in {
    'Aïssata Kébé': 55000, 'Maryam Mamadou BA': 55000,
    'Sokhna Mariama Bousso BADIANE': 55000, 'Mouhamed Salih BADIANE': 55000,
    'El Hadji Mouhamadou Moustapha NGOM': 55000, 'Papa Sélé Mbaye': 55000,
    'Fanta MBAYE': 55000, 'Keba FALL': 55000, 'Adjia Fatou TRAORE': 55000,
    'Fatou Kiné TRAORE': 55000, 'Khady Mountakha DIOP': 55000,
    'Abdoulaye NDIAYE': 27500, 'Mouhamed Abdallah NDIAYE': 27500,
    'Diariatoulah TALL': 32500,
}.items()}

# Nouveaux élèves : inscription 185 000 au lieu du renouvellement (reste dû).
NOUVEAUX_INSCRIPTION = {_norm(n): Decimal(str(v)) for n, v in {
    'Goundo Momi Keita': 0, 'Aïssatou Keïta': 0, 'Mamy Daya CISSOKHO': 85000,
}.items()}


class Command(BaseCommand):
    help = "Reconstruit le reste dû réel par élève (modèle Shoumoul)."

    def add_arguments(self, parser):
        parser.add_argument('--fichier', required=True,
                            help="Excel élèves (colonnes À jour / Dette actuelle).")
        parser.add_argument('--tenant_id')
        parser.add_argument('--exercice')
        parser.add_argument('--appliquer', action='store_true')

    def handle(self, *args, **o):
        from apps.eleves.models import Eleve
        tenant = self._tenant(o.get('tenant_id'))
        ex = self._exercice(tenant, o.get('exercice'))
        statuts = self._lire_fichier(o['fichier'])   # {norm_nom: (a_jour bool, dette Decimal)}

        self.stdout.write(self.style.MIGRATE_HEADING(
            f"\n═══ Reconstruction reste dû — {tenant.nom} — {ex.annee_scolaire} ═══\n"))

        eleves = list(Eleve.objects.filter(tenant=tenant, exercice=ex)
                      .select_related('section').prefetch_related('abonnements__service'))
        plan, alertes = [], []
        total_reste = Decimal('0')
        vus = set()

        for e in eleves:
            if not e.section:
                alertes.append(f"{e.nom_complet} : sans section, ignoré"); continue
            key = _norm(e.nom_complet)
            st = statuts.get(key)
            if st is None:
                # Ajouté au système après le fichier → traité à jour, renouvellement dû.
                a_jour, dette = True, Decimal('0')
                alertes.append(f"{e.nom_complet} : absent du fichier → traité à jour, renouvellement dû")
            else:
                vus.add(key)
                a_jour, dette = st

            M = Decimal(str(e.frais_mensualite_effectif))
            nb_dues = e.nb_mensualites_dues or 1
            serv_mensuel = Decimal(str(e.montant_services_annuel)) / nb_dues
            mensuel = M + serv_mensuel
            attendu = Decimal(str(e.total_attendu))
            verse = RENOUV_VERSE.get(key, Decimal('0'))

            if key in NOUVEAUX_INSCRIPTION:
                # Nouveaux : inscription au lieu du renouvellement. À jour → 5 mois +
                # inscription due ; non à jour → l'inscription due EST le reste (Mamy Daya 85 000).
                insc = NOUVEAUX_INSCRIPTION[key]
                reste = (MOIS_RESTANTS * mensuel + insc) if a_jour else insc
            elif not a_jour:
                reste = dette                                # N → dette = reste complet
            else:
                reste = MOIS_RESTANTS * mensuel + (RENOUVELLEMENT - verse)  # O → 5 mois + renouv
            reste = max(min(reste, attendu), Decimal('0'))   # borné [0, attendu]
            paye = attendu - reste
            plan.append((e, mensuel, verse, a_jour, dette, paye, reste, attendu))
            total_reste += reste

        for key in statuts:
            if key not in vus:
                alertes.append(f"Fichier : « {key} » non trouvé dans la base")
        for key in list(RENOUV_VERSE) + list(NOUVEAUX_INSCRIPTION):
            if key not in vus:
                alertes.append(f"Liste renouvellement/inscription : « {key} » non apparié")

        self.stdout.write("  Élève                               mensuel  renouv àjour   reste dû")
        for e, mensuel, verse, a_jour, dette, paye, reste, attendu in plan:
            self.stdout.write(f"    {e.nom_complet[:33]:<33} {mensuel:>8,.0f} {verse:>7,.0f} "
                              f"{'O' if a_jour else 'N':>4}  {reste:>11,.0f}")
        self.stdout.write(self.style.MIGRATE_LABEL(
            f"\n  {len(plan)} élèves · reste dû total {total_reste:,.0f}"))
        if alertes:
            self.stdout.write(self.style.WARNING("\n  ⚠ À vérifier :"))
            for a in alertes:
                self.stdout.write(self.style.WARNING(f"    - {a}"))

        if not o['appliquer']:
            self.stdout.write(self.style.MIGRATE_LABEL(
                "\n  DRY-RUN — aucune modification. --appliquer pour reconstruire."))
            return

        with transaction.atomic():
            anciennes = Paiement.objects.filter(tenant=tenant, exercice=ex, mode_paiement='REPRISE')
            ids = list(anciennes.values_list('id', flat=True))
            JournalEntry.objects.filter(tenant=tenant, exercice=ex, source='PAIEMENT',
                                        source_id__in=ids).delete()
            JournalEntry.objects.filter(tenant=tenant, exercice=ex, no_piece='RECAL-REP').delete()
            anciennes.delete()

            for e, mensuel, verse, a_jour, dette, paye, reste, attendu in plan:
                if paye <= 0:
                    continue
                M = Decimal(str(e.frais_mensualite_effectif))
                mens = min(MOIS_PAYES * M, paye)             # mensualités « payées »
                montants = {
                    'montant_inscription': 0,
                    'montant_mensualite':  float(mens),
                    'montant_uniforme':    0,
                    'montant_fournitures': 0,
                    'montant_divers':      float(paye - mens),  # services + inscription absorbés
                }
                # creer_paiement_reprise n'accepte pas montant_divers → on l'injecte après.
                p = creer_paiement_reprise(tenant, ex, e, montants={
                    'montant_inscription': 0, 'montant_mensualite': float(mens),
                    'montant_uniforme': 0, 'montant_fournitures': 0})
                if p and montants['montant_divers'] > 0:
                    p.montant_divers = Decimal(str(montants['montant_divers']))
                    p.save(update_fields=['montant_divers'])
                    # ajuste les écritures (créance + produit) du delta divers
                    delta = Decimal(str(montants['montant_divers']))
                    JournalEntry.objects.filter(tenant=tenant, exercice=ex, source='PAIEMENT',
                                                source_id=p.id, no_compte='411', debit__gt=0).update(
                        debit=Decimal(str(mens)) + delta)
                    JournalEntry.objects.filter(tenant=tenant, exercice=ex, source='PAIEMENT',
                                                source_id=p.id, no_compte='706').update(
                        credit=Decimal(str(mens)) + delta)
                    JournalEntry.objects.filter(tenant=tenant, exercice=ex, source='PAIEMENT',
                                                source_id=p.id, no_compte='890').update(
                        debit=Decimal(str(mens)) + delta)
                    JournalEntry.objects.filter(tenant=tenant, exercice=ex, source='PAIEMENT',
                                                source_id=p.id, no_compte='411', credit__gt=0).update(
                        credit=Decimal(str(mens)) + delta)

            neuf = list(Paiement.objects.filter(tenant=tenant, exercice=ex,
                                                mode_paiement='REPRISE').values_list('id', flat=True))
            r706 = Decimal(str(JournalEntry.objects.filter(
                tenant=tenant, exercice=ex, source='PAIEMENT', source_id__in=neuf,
                no_compte='706', credit__gt=0).aggregate(c=Sum('credit'))['c'] or 0))
            if r706 > 0:
                JournalEntry.objects.bulk_create([
                    JournalEntry(tenant=tenant, exercice=ex, no_piece='RECAL-REP',
                                 date_ecriture=ex.date_debut, source='RECAL_MIGRATION',
                                 no_compte='706', debit=r706, credit=0, ordre=1,
                                 libelle="Neutralisation reprise reconstruite (706 = agrégats Excel)"),
                    JournalEntry(tenant=tenant, exercice=ex, no_piece='RECAL-REP',
                                 date_ecriture=ex.date_debut, source='RECAL_MIGRATION',
                                 no_compte='890', debit=0, credit=r706, ordre=2,
                                 libelle="Contrepartie neutralisation reprise reconstruite"),
                ])

        self.stdout.write(self.style.SUCCESS(
            f"\n  ✓ Reconstruit : {len(plan)} élèves · reste dû total {total_reste:,.0f}. "
            f"706 inchangé (reprises neutralisées en 890)."))

    def _lire_fichier(self, chemin):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(chemin, data_only=True, read_only=True)
        except Exception as e:
            raise CommandError(f"Fichier illisible : {e}")
        ws = wb['Élèves'] if 'Élèves' in wb.sheetnames else wb.active
        out = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or not row[0]:
                continue
            nom = row[0]
            a_jour = str(row[17] if len(row) > 17 else '').strip().upper() != 'N'
            dette = row[18] if len(row) > 18 else None
            dette = Decimal(str(dette)) if dette not in (None, '') else Decimal('0')
            out[_norm(nom)] = (a_jour, dette)
        return out

    def _tenant(self, tid):
        if tid:
            try:
                return Tenant.objects.get(id=tid)
            except Tenant.DoesNotExist:
                raise CommandError(f'Tenant {tid} introuvable')
        ts = list(Tenant.objects.all()[:2])
        if len(ts) == 1:
            return ts[0]
        raise CommandError('Plusieurs tenants : préciser --tenant_id')

    def _exercice(self, tenant, annee):
        qs = Exercice.objects.filter(tenant=tenant)
        ex = qs.filter(annee_scolaire=annee).first() if annee else \
            qs.filter(cloture=False).order_by('-date_debut').first()
        if not ex:
            raise CommandError("Exercice introuvable : préciser --exercice")
        return ex
