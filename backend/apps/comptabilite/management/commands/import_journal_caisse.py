"""
Commande : python manage.py import_journal_caisse <fichier.xlsx> [--tenant_id UUID]
                                                  [--dry-run] [--feuilles 2021,2022]

Migration de l'historique d'un journal de caisse Excel (écoles reprises)
vers la comptabilité SYSCOHADA de SAGI SCHOOL.

Format attendu (une feuille par année civile, nommée « 2021 », « 2022 », …) :
    DATE | N°réçu | Description | RUBRIQUE | SOURCE | DESTINATION |
    N°.Facture | ENTREE | SORTIE | SOLDE | commentaire
La ligne d'en-tête est détectée automatiquement (colonnes DATE + ENTREE),
les dates vides reprennent la dernière date vue, les feuilles vides sont
ignorées.

Règles de ventilation (démarche HADY GESMAN : entrées élèves = produits 706,
le reste ventilé selon sa nature) :
  - ENTREE  : scolarité / inscription / mensualité → 706 (défaut entrée),
              uniformes → 707, sport/activités → 706.3, dons → 758,
              prêts/remboursements → 46.
  - SORTIE  : ventilée par rubrique puis par mots-clés de la description
              (restauration → 604, salaires → 661, loyer → 622, eau → 6051,
              électricité → 6052, téléphone → 628, …), défaut → 658.
  - Chaque mouvement génère 2 lignes équilibrées avec la caisse (571) en
    contrepartie : entrée = 571 D / compte C ; sortie = compte D / 571 C.
  - Une ligne « SOLDE » en début de feuille alimente
    exercice.solde_initial_caisse (jamais d'écriture : les tableaux de bord
    additionnent déjà ce champ au net des écritures 571).
  - Les exercices historiques sont créés (année civile) et clôturés.

Le rapport final compare le solde de caisse recalculé au dernier SOLDE de la
feuille : tout écart signale des lignes incohérentes dans l'Excel d'origine
(fréquent — la plupart de ces établissements tenaient leur caisse à la main).
"""
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal

import calendar

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from apps.comptabilite.models import JournalEntry
from apps.paiements.models import Exercice
from apps.tenants.models import Tenant

CAISSE = '571'

# Rubriques normalisées (sans accents, majuscules) → (compte, sens)
# sens : 'E' = ne s'applique qu'aux entrées, 'S' = sorties, None = les deux.
MAPPING_RUBRIQUES = [
    # ── Entrées élèves → produits ──
    (r'SCOLARIT|INSCRIPTION|MENSUALIT|REINSCRIPTION', '706',   'E'),
    (r'UNIFORME',                                     '707',   'E'),
    (r'SPORT|EXTRASCOLAIRE|ACTIVIT',                  '706.3', 'E'),
    (r'FOURNITURE',                                   '706',   'E'),
    (r'CANTINE|RESTAURATION|GOUTER',                  '706.1', 'E'),
    (r'DON|SUBVENTION',                               '758',   'E'),
    (r'PRET|REMBOURSEMENT',                           '46',    None),
    (r'DIVERS',                                       '758',   'E'),
    # ── Sorties → charges par nature ──
    (r'RESTAURATION|GOUTER|RAVITAILLEMENT|BOUTIQUE|BOULANGERIE|MARCHE|VIANDE|POISSON|POULET|ARRAW|INTENDAN|LEGUME|\bGAZ\b|CHARBON|\bRIZ\b|HUILE', '604', 'S'),
    (r'SALAIRE|MOTIVATION|PERSONNEL|OUSTAZ|\bTATA',   '661',  'S'),
    (r'LOYER|LOCATION',                               '622',  'S'),
    (r'ELECTRICITE|W?OYOFAL|SENELEC',                 '6052', 'S'),
    (r"SEN.?'?EAU|\bEAU\b",                           '6051', 'S'),
    (r'SONATEL|COMMUNICATION|CONNEXION|TELEPHONE|INTERNET|WIFI', '628', 'S'),
    (r'FOURNITURE|MATERIEL|MATELAS|TAPIS|MOQUETTE|HOUSSE|CONGELATEUR|FRIGO|MACHINE|CHAISE|TABLE|LIVRE', '6054', 'S'),
    (r'ENTRETIEN|MAINTENANCE|REPAR|TOITURE|CARRELAGE|VID.?NGE|DESINFECT|PEINTURE|NETTOYAGE', '624',  'S'),
    (r'S.?ECURIT|GARDIEN',                            '621',  'S'),
    (r'TRANSP|DEPLACEMENT|DEMENAGEMENT',              '618',  'S'),
    (r'SANTE|PHARMACIE|MEDICAMENT|MEDECIN',           '658',  'S'),
    (r'AUTORISATION|NINEA|RCCM|DEPOT.?DOSSIER|TIMBRE|ENREGISTREMENT', '645', 'S'),
    (r'ASSURANCE',                                    '625',  'S'),
    (r'FORMATION',                                    '633',  'S'),
    (r'BANQUE|FRAIS.?BANC',                           '631',  'S'),
    (r'DETTE',                                        '16',   None),
    (r'PROJET',                                       '658',  'S'),
    (r'LINGE|SAVON|BALAI',                            '605',  'S'),
]

# Défauts quand ni la rubrique ni la description ne matchent.
DEFAUT_ENTREE = '706'   # démarche : toute entrée non identifiée = produit scolarité
DEFAUT_SORTIE = '658'   # charges diverses

MOIS_FR = {
    'JANVIER': 1, 'FEVRIER': 2, 'MARS': 3, 'AVRIL': 4, 'MAI': 5, 'JUIN': 6,
    'JUILLET': 7, 'AOUT': 8, 'SEPTEMBRE': 9, 'OCTOBRE': 10,
    'NOVEMBRE': 11, 'DECEMBRE': 12,
}


def _norm(texte):
    """Majuscules sans accents pour un matching tolérant (RUBLIQUE, Dêtte…)."""
    if texte is None:
        return ''
    s = unicodedata.normalize('NFKD', str(texte))
    return ''.join(c for c in s if not unicodedata.combining(c)).upper().strip()


def _montant(valeur):
    if valeur is None:
        return None
    if isinstance(valeur, (int, float)):
        return Decimal(str(round(float(valeur), 2)))
    txt = str(valeur).replace(' ', '').replace(' ', '').replace(',', '.')
    try:
        return Decimal(txt)
    except Exception:
        return None


def _parse_date(valeur, annee, derniere):
    """Date de la cellule ; texte type '01-05/01/2023' → dernier jj/mm/aaaa ;
    mois en toutes lettres → fin de mois ; sinon reprend la dernière date."""
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    if valeur is not None:
        txt = _norm(valeur)
        m = re.findall(r'(\d{1,2})/(\d{1,2})/(\d{4})', txt)
        if m:
            j, mo, a = map(int, m[-1])
            try:
                return date(a, mo, j)
            except ValueError:
                pass
        for nom, num in MOIS_FR.items():
            if nom in txt:
                return date(annee, num, calendar.monthrange(annee, num)[1])
    return derniere


def _date_depuis_description(description, annee):
    """« Total Entrées Janvier » → 31/01/annee (feuilles agrégées type 2024)."""
    txt = _norm(description)
    for nom, num in MOIS_FR.items():
        if nom in txt:
            return date(annee, num, calendar.monthrange(annee, num)[1])
    return None


def _mapper(rubrique, description, sens):
    """Rend (no_compte, via) — via = 'rubrique', 'description' ou 'defaut'."""
    for source, texte in (('rubrique', _norm(rubrique)),
                          ('description', _norm(description))):
        if not texte:
            continue
        for motif, compte, s in MAPPING_RUBRIQUES:
            if s is not None and s != sens:
                continue
            if re.search(motif, texte):
                return compte, source
    return (DEFAUT_ENTREE if sens == 'E' else DEFAUT_SORTIE), 'defaut'


class Command(BaseCommand):
    help = "Importe l'historique d'un journal de caisse Excel en écritures SYSCOHADA"

    def add_arguments(self, parser):
        parser.add_argument('fichier', help='Chemin du fichier .xlsx')
        parser.add_argument('--tenant_id', help='UUID du tenant (facultatif si un seul)')
        parser.add_argument('--feuilles', help='Années à importer, ex. 2021,2022 (défaut : toutes)')
        parser.add_argument('--dry-run', action='store_true',
                            help='Analyse et rapport sans rien écrire en base')
        parser.add_argument('--sans-cloture', action='store_true',
                            help='Ne pas clôturer les exercices créés')

    # ------------------------------------------------------------------ #
    def handle(self, *args, **o):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl n'est pas installé sur ce poste.")

        tenant = self._tenant(o.get('tenant_id'))
        wb = openpyxl.load_workbook(o['fichier'], data_only=True)

        voulues = None
        if o.get('feuilles'):
            voulues = {a.strip() for a in o['feuilles'].split(',')}

        feuilles = [n for n in wb.sheetnames if re.fullmatch(r'\d{4}', n.strip())
                    and (voulues is None or n.strip() in voulues)]
        if not feuilles:
            raise CommandError("Aucune feuille annuelle (nommée « 2021 », « 2022 », …) trouvée.")

        ignorees = [n for n in wb.sheetnames if n not in feuilles]
        if ignorees:
            self.stdout.write(f"Feuilles ignorées : {', '.join(ignorees)}")

        with transaction.atomic():
            for nom in feuilles:
                self._importer_feuille(tenant, wb[nom], int(nom.strip()), o)
            if o['dry_run']:
                self.stdout.write(self.style.WARNING(
                    '\nDRY-RUN : aucune donnée écrite (transaction annulée).'))
                transaction.set_rollback(True)

    # ------------------------------------------------------------------ #
    def _tenant(self, tenant_id):
        if tenant_id:
            try:
                return Tenant.objects.get(id=tenant_id)
            except Tenant.DoesNotExist:
                raise CommandError(f'Tenant {tenant_id} introuvable')
        tenants = list(Tenant.objects.all()[:2])
        if len(tenants) == 1:
            return tenants[0]
        raise CommandError('Plusieurs tenants : préciser --tenant_id UUID')

    # ------------------------------------------------------------------ #
    def _importer_feuille(self, tenant, ws, annee, o):
        lignes = list(ws.iter_rows(values_only=True))
        entete = self._trouver_entete(lignes)
        if entete is None:
            self.stdout.write(self.style.WARNING(
                f'{annee} : en-tête introuvable, feuille ignorée.'))
            return

        idx, colonnes = entete
        mouvements, solde_initial, dernier_solde, anomalies = \
            self._parser(lignes[idx + 1:], colonnes, annee)

        if not mouvements:
            self.stdout.write(f'{annee} : aucune donnée, feuille ignorée.')
            return

        exercice = None
        if not o['dry_run']:
            exercice = self._exercice(tenant, annee, solde_initial, o)
            if exercice is None:
                return

        # Écritures
        tot_e = tot_s = Decimal(0)
        par_compte, par_defaut = {}, {}
        for seq, mvt in enumerate(mouvements, 1):
            compte, via = _mapper(mvt['rubrique'], mvt['description'], mvt['sens'])
            libelle = mvt['description'] or mvt['rubrique'] or 'Mouvement de caisse'
            if mvt['rubrique'] and mvt['rubrique'] not in libelle:
                libelle = f"{libelle} [{mvt['rubrique']}]"
            no_piece = f"MIG{annee % 100:02d}-{mvt['recu'] or f'{seq:04d}'}"[:30]

            if mvt['sens'] == 'E':
                ecritures = [(1, CAISSE, mvt['montant'], 0), (2, compte, 0, mvt['montant'])]
                tot_e += mvt['montant']
            else:
                ecritures = [(1, compte, mvt['montant'], 0), (2, CAISSE, 0, mvt['montant'])]
                tot_s += mvt['montant']

            cle = (mvt['sens'], compte)
            par_compte[cle] = par_compte.get(cle, Decimal(0)) + mvt['montant']
            if via == 'defaut':
                par_defaut[cle] = par_defaut.get(cle, Decimal(0)) + mvt['montant']

            if not o['dry_run']:
                for ordre, no_compte, debit, credit in ecritures:
                    JournalEntry.objects.create(
                        tenant=tenant, exercice=exercice,
                        no_piece=no_piece, date_ecriture=mvt['date'],
                        no_compte=no_compte, libelle=f'Migration — {libelle}'[:500],
                        debit=debit, credit=credit,
                        source='MIGRATION', ordre=ordre,
                    )

        # Rapport
        solde_ini = solde_initial if solde_initial is not None else Decimal(0)
        self.stdout.write(self.style.MIGRATE_HEADING(f'\n═══ Exercice {annee} ═══'))
        self.stdout.write(f'  {len(mouvements)} mouvements | entrées {tot_e:,.0f} | '
                          f'sorties {tot_s:,.0f} | solde initial {solde_ini:,.0f}'
                          + ('' if solde_initial is not None else ' (pas de ligne SOLDE)'))
        calcule = solde_ini + tot_e - tot_s
        self.stdout.write(f'  Solde caisse recalculé : {calcule:,.0f}'
                          + (f' | dernier SOLDE Excel : {dernier_solde:,.0f}'
                             if dernier_solde is not None else ''))
        if dernier_solde is not None and abs(calcule - dernier_solde) > Decimal('0.5'):
            self.stdout.write(self.style.WARNING(
                f'  ⚠ Écart de {calcule - dernier_solde:,.0f} avec le SOLDE Excel '
                '(lignes incohérentes dans le fichier source — à vérifier avec le directeur).'))
        for (sens, compte), total in sorted(par_compte.items()):
            marque = ' (défaut)' if (sens, compte) in par_defaut else ''
            self.stdout.write(f"    {'ENTRÉE' if sens == 'E' else 'SORTIE'} "
                              f'{compte:<6} : {total:>12,.0f}{marque}')
        for a in anomalies:
            self.stdout.write(self.style.WARNING(f'  ⚠ {a}'))

    # ------------------------------------------------------------------ #
    def _trouver_entete(self, lignes):
        for i, ligne in enumerate(lignes[:15]):
            cellules = [_norm(c) for c in ligne]
            if 'DATE' in cellules and any('ENTREE' in c for c in cellules):
                col = {}
                for j, c in enumerate(cellules):
                    if not c:
                        continue
                    if c == 'DATE':                 col['date'] = j
                    elif 'RECU' in c or 'REÇU' in c: col['recu'] = j
                    elif 'DESCRIPTION' in c:         col['description'] = j
                    elif 'RUBRIQUE' in c or 'RUBLIQUE' in c: col['rubrique'] = j
                    elif 'ENTREE' in c:              col['entree'] = j
                    elif 'SORTIE' in c:              col['sortie'] = j
                    elif c == 'SOLDE':               col['solde'] = j
                return i, col
        return None

    # ------------------------------------------------------------------ #
    def _parser(self, lignes, col, annee):
        def val(ligne, cle):
            j = col.get(cle)
            return ligne[j] if j is not None and j < len(ligne) else None

        mouvements, anomalies = [], []
        solde_initial = None      # None = pas de ligne SOLDE dans la feuille
        dernier_solde = None
        derniere_date = date(annee, 1, 1)

        for ligne in lignes:
            entree = _montant(val(ligne, 'entree'))
            sortie = _montant(val(ligne, 'sortie'))
            solde = _montant(val(ligne, 'solde'))
            if solde is not None:
                dernier_solde = solde
            if not entree and not sortie:
                continue

            description = str(val(ligne, 'description') or '').strip()
            rubrique = str(val(ligne, 'rubrique') or '').strip()

            # Ligne de report « SOLDE 2021 » → solde initial de l'exercice
            if re.search(r'\bSOLDE\b', _norm(description) + ' ' + _norm(rubrique)):
                solde_initial = entree or (sortie and -sortie) or Decimal(0)
                continue
            # Ligne de total général (ni description ni rubrique)
            if not description and not rubrique:
                anomalies.append(f'Ligne sans description ignorée '
                                 f'(E={entree or 0:,.0f} / S={sortie or 0:,.0f}) — '
                                 'probablement un total.')
                continue

            d = _parse_date(val(ligne, 'date'), annee, derniere_date)
            d = _date_depuis_description(description, annee) or d
            if d.year != annee:   # report du 31/12 N-1, saisies débordantes…
                d = date(annee, 1, 1) if d.year < annee else date(annee, 12, 31)
            derniere_date = d

            recu = str(val(ligne, 'recu') or '').strip() or None
            # Certaines lignes portent à la fois une entrée ET une sortie.
            if entree:
                mouvements.append(dict(sens='E', montant=entree, date=d,
                                       recu=recu, description=description,
                                       rubrique=rubrique))
            if sortie:
                mouvements.append(dict(sens='S', montant=sortie, date=d,
                                       recu=recu, description=description,
                                       rubrique=rubrique))
        return mouvements, solde_initial, dernier_solde, anomalies

    # ------------------------------------------------------------------ #
    def _exercice(self, tenant, annee, solde_initial, o):
        exercice, cree = Exercice.objects.get_or_create(
            tenant=tenant, annee_scolaire=str(annee),
            defaults=dict(
                date_debut=date(annee, 1, 1), date_fin=date(annee, 12, 31),
                nb_mensualites=12, solde_initial_caisse=solde_initial or 0,
            ),
        )
        if not cree:
            deja = JournalEntry.objects.filter(
                tenant=tenant, exercice=exercice, source='MIGRATION').count()
            if deja:
                self.stdout.write(self.style.WARNING(
                    f'{annee} : {deja} écritures MIGRATION déjà présentes — '
                    'feuille ignorée (relancer après les avoir supprimées si besoin).'))
                return None
            # Ne pas écraser le solde saisi sur un exercice existant (ex. le
            # comptage physique de l'exercice courant) sans ligne SOLDE.
            if solde_initial is not None:
                exercice.solde_initial_caisse = solde_initial
        if not o['sans_cloture']:
            exercice.cloture = True
            exercice.date_cloture = exercice.date_cloture or timezone.now()
        exercice.save()
        return exercice
