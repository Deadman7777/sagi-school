"""Import Excel des charges (dépenses) — migration/saisie en masse.

Même esprit que l'import des élèves : on télécharge un modèle, on téléverse le
fichier rempli, on obtient un aperçu ligne par ligne (statut OK / ERREUR, compte
suggéré par nature), puis on confirme pour générer les écritures SYSCOHADA.

Chaque ligne OK → 2 écritures équilibrées : `6xx D (charge) / trésorerie C`
(571 caisse par défaut). Source 'CHARGE' pour apparaître dans la liste des
charges. openpyxl est importé paresseusement (poste Windows sans la lib).
"""
import re
import datetime
from decimal import Decimal, InvalidOperation

# Suggestion de compte de charge (6xx) d'après le libellé — sous-ensemble
# « sorties » du mapping du journal de caisse.
MAPPING_CHARGES = [
    (r'RESTAURATION|GOUTER|RAVITAILLEMENT|BOUTIQUE|BOULANGERIE|MARCHE|VIANDE|POISSON|POULET|INTENDAN|LEGUME|\bGAZ\b|CHARBON|\bRIZ\b|HUILE', '604'),
    (r'SALAIRE|MOTIVATION|PERSONNEL|OUSTAZ|\bTATA', '661'),
    (r'LOYER|LOCATION', '622'),
    (r'ELECTRICITE|WOYOFAL|SENELEC', '6052'),
    (r"SEN.?'?EAU|\bEAU\b", '6051'),
    (r'SONATEL|COMMUNICATION|CONNEXION|TELEPHONE|INTERNET|WIFI', '628'),
    (r'FOURNITURE|MATERIEL|MATELAS|TAPIS|MOQUETTE|CONGELATEUR|FRIGO|MACHINE|CHAISE|TABLE|LIVRE', '6054'),
    (r'ENTRETIEN|MAINTENANCE|REPAR|TOITURE|CARRELAGE|VIDANGE|DESINFECT|PEINTURE|NETTOYAGE', '624'),
    (r'SECURIT|GARDIEN', '621'),
    (r'TRANSP|DEPLACEMENT|DEMENAGEMENT', '618'),
    (r'SANTE|PHARMACIE|MEDICAMENT|MEDECIN', '658'),
    (r'AUTORISATION|NINEA|RCCM|TIMBRE|ENREGISTREMENT', '645'),
    (r'ASSURANCE', '625'),
    (r'FORMATION', '633'),
    (r'BANQUE|FRAIS.?BANC', '631'),
    (r'LINGE|SAVON|BALAI', '605'),
]
DEFAUT_CHARGE = '658'  # charges diverses
COMPTES_TRESORERIE = {'571', '5715', '521', '5521', '5522', '5523'}

COLONNES = ['Date', 'Libellé', 'Compte (optionnel)', 'Montant', 'Réglé via (571 défaut)']


def _openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError(
            "La bibliothèque openpyxl n'est pas installée sur ce poste. "
            "Exécutez « python -m pip install openpyxl » puis relancez l'application.")


def _norm(s):
    import unicodedata
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode()
    return s.upper().strip()


def suggerer_compte(libelle):
    txt = _norm(libelle)
    for motif, compte in MAPPING_CHARGES:
        if re.search(motif, txt):
            return compte
    return DEFAUT_CHARGE


def _date(val):
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.date() if isinstance(val, datetime.datetime) else val
    s = str(val or '').strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _montant(val):
    if val in (None, ''):
        return None
    try:
        return Decimal(str(val).replace(' ', '').replace(' ', '').replace(',', '.'))
    except (InvalidOperation, ValueError):
        return None


def generer_template(tenant):
    openpyxl = _openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Charges'
    ws.append(COLONNES)
    # Exemples indicatifs
    ws.append(['15/01/2026', 'Loyer janvier',        '',    150000, '571'])
    ws.append(['20/01/2026', 'Facture SENELEC',       '6052', 45000, '571'])
    ws.append(['25/01/2026', 'Salaire oustaz Modou',  '',    120000, '571'])
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def analyser(fichier, tenant, exercice):
    """Rend {'resume': {...}, 'lignes': [...]}. Chaque ligne :
    {ligne, date, libelle, no_compte, compte_suggere(bool), montant,
     compte_tresorerie, statut OK|ERREUR, erreurs[]}."""
    openpyxl = _openpyxl()
    try:
        wb = openpyxl.load_workbook(fichier, read_only=True, data_only=True)
    except Exception:
        raise ValueError("Fichier illisible — envoyez un .xlsx (Excel 2007+), pas un .xls ni un CSV.")
    ws = wb['Charges'] if 'Charges' in wb.sheetnames else wb.active

    lignes, total = [], Decimal(0)
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) > 5000:
        raise ValueError("Fichier trop volumineux (plus de 5000 lignes).")

    def _txt(row, i):
        v = row[i] if len(row) > i else None
        return '' if v is None else str(v).strip()

    for i, row in enumerate(rows[1:], start=2):  # saute l'en-tête
        if row is None or all(c in (None, '') for c in row):
            continue
        date_v = _date(row[0] if len(row) > 0 else None)
        libelle = _txt(row, 1)
        compte_saisi = _txt(row, 2)
        montant = _montant(row[3] if len(row) > 3 else None)
        tresor = _txt(row, 4) or '571'

        erreurs = []
        if not libelle:
            erreurs.append('Libellé manquant')
        if montant is None or montant <= 0:
            erreurs.append('Montant invalide')
        if date_v is None:
            erreurs.append('Date invalide (jj/mm/aaaa)')
        if tresor not in COMPTES_TRESORERIE:
            tresor = '571'
        compte = compte_saisi or suggerer_compte(libelle)
        if not compte.startswith('6'):
            erreurs.append(f'Compte {compte} : une charge doit être en classe 6')

        lignes.append({
            'ligne': i,
            'date': date_v.isoformat() if date_v else '',
            'libelle': libelle,
            'no_compte': compte,
            'compte_suggere': not compte_saisi and not erreurs,
            'montant': float(montant) if montant else 0,
            'compte_tresorerie': tresor,
            'statut': 'OK' if not erreurs else 'ERREUR',
            'erreurs': erreurs,
        })
        if not erreurs:
            total += montant

    ok = sum(1 for l in lignes if l['statut'] == 'OK')
    return {
        'resume': {
            'total_lignes': len(lignes),
            'ok': ok,
            'erreurs': len(lignes) - ok,
            'montant_total': float(total),
            'exercice': exercice.annee_scolaire,
        },
        'lignes': lignes,
    }
