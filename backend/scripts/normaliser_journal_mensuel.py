"""
Normalise un journal de caisse « une feuille par mois » vers le format
canonique attendu par `manage.py import_journal_caisse` (une feuille par
année : DATE | N°réçu | Description | RUBRIQUE | … | ENTREE | SORTIE).

Usage :
    python normaliser_journal_mensuel.py "JOURNAL CAISSE 2025_ A jour.xlsx" \
           [sortie.xlsx]

Format d'entrée (relevé chez les écoles qui tiennent un classeur par an) :
  - feuilles « JANVIER 25 », « FEVRIER 25 », … « MAI 2025 » : en-tête
    DATES | LIBELLE | ENTREES | SORTIES quelque part dans les 15 premières
    lignes ; lignes « SOLDE <mois> » (cumuls) et « TOTAUX » à ignorer.
  - feuille « RECAP <année> » facultative : MOIS | LIBELLE | RECETTES |
    DEPENSES. Les mois du RECAP sans feuille de détail sont repris en
    2 mouvements agrégés (entrées / sorties) datés fin de mois — et le
    RECAP sert de contrôle des totaux.

Le fichier produit s'importe ensuite sur le poste de l'école :
    python manage.py import_journal_caisse sortie.xlsx --feuilles 2025 ...
"""
import re
import sys
import calendar
import unicodedata
from datetime import date, datetime
from decimal import Decimal

import openpyxl

MOIS_FR = {
    'JANVIER': 1, 'FEVRIER': 2, 'MARS': 3, 'AVRIL': 4, 'MAI': 5, 'JUIN': 6,
    'JUILLET': 7, 'AOUT': 8, 'SEPTEMBRE': 9, 'OCTOBRE': 10,
    'NOVEMBRE': 11, 'DECEMBRE': 12,
}

ENTETE_CANONIQUE = ['DATE', 'N°réçu', 'Description', 'RUBRIQUE', 'SOURCE',
                    'DESTINATION', 'N°.Facture', 'ENTREE', 'SORTIE', 'SOLDE',
                    'commentaire']

# Indices (rubrique steering) pour les sorties que le mapping par mots-clés
# de import_journal_caisse ne couvre pas dans le bon sens. Évalués dans
# l'ordre, premier motif gagnant ; rubrique None = ne rien poser (laisser la
# commande décider sur la description) — sert à protéger l'alimentation des
# motifs plus larges placés en dessous (ex. COMMANDE de viande ≠ mobilier).
RUBRIQUES_INDICES = [
    (r'RAVITAILLEMENT|BOULANGERIE|MARCHE|LEGUME|VIANDE|POULET|POISSON|OIGNON|\bPATE|\bBOLS?\b|GOUTER', None),
    (r'UNIFORME',                                  'MATERIELS'),           # achat d'uniformes → 6054
    (r'\bCREDIT\b',                                'COMMUNICATION'),       # crédit téléphone → 628
    # Personnel : avances sur salaires et paiements nominatifs (tatas, oustaz…)
    (r'AVCE|AV/SAL|/SAL\b|\bSAL\b|^TA\s|^TATA|AVANCE\s+TA\b|OUS\.?\s*NIANG|BABOU|AISSATOU DIOP', 'SALAIRE / MOTIVATION'),
    (r'COMMANDE|MENUISIER|ETAGERE|SERVIETTE|FOURNEAU|ONDULEUR|TIRELIRE',   'MATERIELS'),
    (r'HOSPITAL|HOPITAL|CLINIQUE|URGENCE',         'SANTE'),
    (r'FORMALITE|FRAIS DE DOSSIER|TERRAIN',        'AUTORISATION'),        # → 645
]


def norm(t):
    if t is None:
        return ''
    s = unicodedata.normalize('NFKD', str(t))
    return ''.join(c for c in s if not unicodedata.combining(c)).upper().strip()


def montant(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(round(float(v), 2))) or None
    txt = str(v).replace(' ', '').replace(' ', '').replace(',', '.')
    try:
        return Decimal(txt) or None
    except Exception:
        return None


def mois_de_feuille(nom):
    """« JANVIER 25 » → (2025, 1) ; « MAI 2025 » → (2025, 5) ; sinon None."""
    t = norm(nom)
    for nom_mois, num in MOIS_FR.items():
        if t.startswith(nom_mois):
            m = re.search(r'(\d{2,4})\s*$', t)
            if m:
                annee = int(m.group(1))
                if annee < 100:
                    annee += 2000
                return annee, num
    return None


def parser_feuille_mois(ws, annee, mois):
    """Rend la liste des mouvements [(date, libelle, rubrique, e, s)]."""
    lignes = list(ws.iter_rows(values_only=True))
    entete = None
    for i, l in enumerate(lignes[:15]):
        cells = [norm(c) for c in l]
        if any(c.startswith('DATE') for c in cells) and any(c.startswith('ENTREE') for c in cells):
            entete = i, {
                'date':   next(j for j, c in enumerate(cells) if c.startswith('DATE')),
                'libelle': next(j for j, c in enumerate(cells) if 'LIBELLE' in c or 'DESCRIPTION' in c),
                'entree': next(j for j, c in enumerate(cells) if c.startswith('ENTREE')),
                'sortie': next(j for j, c in enumerate(cells) if c.startswith('SORTIE')),
            }
            break
    if entete is None:
        print(f'  ⚠ {ws.title} : en-tête introuvable, feuille ignorée')
        return []

    idx, col = entete
    mouvements = []
    derniere = date(annee, mois, 1)
    for l in lignes[idx + 1:]:
        def v(cle):
            j = col[cle]
            return l[j] if j < len(l) else None
        e, s = montant(v('entree')), montant(v('sortie'))
        libelle = str(v('libelle') or '').strip()
        t = norm(libelle) + ' ' + norm(v('entree'))
        if 'SOLDE' in t or 'TOTAUX' in t or 'TOTAL' in t:
            continue                      # cumuls / reports mensuels
        if (e is None and s is None) or not libelle:
            continue
        d = v('date')
        d = d.date() if isinstance(d, datetime) else (d if isinstance(d, date) else derniere)
        if (d.year, d.month) != (annee, mois):      # cellule fantaisiste
            d = derniere
        derniere = d

        rubrique = ''
        if s:   # steering uniquement sur les sorties
            for motif, rub in RUBRIQUES_INDICES:
                if re.search(motif, norm(libelle)):
                    rubrique = rub or ''
                    break
        mouvements.append((d, libelle, rubrique, e, s))
    return mouvements


def parser_recap(wb):
    """{(annee, mois): (recettes, depenses)} depuis les feuilles RECAP <année>."""
    recap = {}
    for nom in wb.sheetnames:
        m = re.match(r'RECAP\s*(\d{4})', norm(nom))
        if not m:
            continue
        annee = int(m.group(1))
        for l in wb[nom].iter_rows(values_only=True):
            mois = MOIS_FR.get(norm(l[0]) if l else '')
            if mois and len(l) > 3:
                r, d = montant(l[2]), montant(l[3])
                if r is not None or d is not None:
                    recap[(annee, mois)] = (r or Decimal(0), d or Decimal(0))
    return recap


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    source = sys.argv[1]
    cible = sys.argv[2] if len(sys.argv) > 2 else re.sub(r'\.xlsx$', '', source) + ' — normalisé.xlsx'

    wb = openpyxl.load_workbook(source, data_only=True)
    par_annee = {}
    for nom in wb.sheetnames:
        am = mois_de_feuille(nom)
        if not am:
            continue
        annee, mois = am
        mvts = parser_feuille_mois(wb[nom], annee, mois)
        par_annee.setdefault(annee, {})[mois] = mvts
        e = sum((m[3] or 0) for m in mvts)
        s = sum((m[4] or 0) for m in mvts)
        print(f'  {nom:<15} : {len(mvts):>3} mouvements | E {e:>12,.0f} | S {s:>12,.0f}')

    recap = parser_recap(wb)
    if not par_annee and not recap:
        sys.exit('Aucune feuille mensuelle (« JANVIER 25 »…) ni RECAP trouvée.')

    out = openpyxl.Workbook()
    out.remove(out.active)
    for annee in sorted(set(list(par_annee) + [a for a, _ in recap])):
        ws = out.create_sheet(str(annee))
        ws.append(ENTETE_CANONIQUE)
        mois_detail = par_annee.get(annee, {})
        tot_e = tot_s = Decimal(0)
        for mois in range(1, 13):
            fin = date(annee, mois, calendar.monthrange(annee, mois)[1])
            if mois in mois_detail:
                for d, libelle, rubrique, e, s in mois_detail[mois]:
                    ws.append([d, '', libelle, rubrique, '', '', '',
                               float(e) if e else None, float(s) if s else None, '', ''])
                    tot_e += e or 0
                    tot_s += s or 0
                # contrôle : détail du mois vs RECAP
                r = recap.get((annee, mois))
                if r:
                    de = sum((m[3] or 0) for m in mois_detail[mois])
                    ds = sum((m[4] or 0) for m in mois_detail[mois])
                    if abs(de - r[0]) > 1 or abs(ds - r[1]) > 1:
                        print(f'  ⚠ {calendar.month_name[mois]} {annee} : détail '
                              f'E {de:,.0f}/S {ds:,.0f} ≠ RECAP E {r[0]:,.0f}/S {r[1]:,.0f}')
            elif (annee, mois) in recap:
                r, d_ = recap[(annee, mois)]
                nom_mois = [k for k, v in MOIS_FR.items() if v == mois][0].capitalize()
                if r:
                    ws.append([fin, '', f'Total entrées {nom_mois} (RECAP — détail manquant)',
                               '', '', '', '', float(r), None, '', ''])
                    tot_e += r
                if d_:
                    ws.append([fin, '', f'Total sorties {nom_mois} (RECAP — détail manquant)',
                               '', '', '', '', None, float(d_), '', ''])
                    tot_s += d_
        print(f'{annee} : TOTAL généré E {tot_e:,.2f} | S {tot_s:,.2f} | net {tot_e - tot_s:,.2f}')

    out.save(cible)
    print(f'→ Fichier normalisé : {cible}')


if __name__ == '__main__':
    main()
